import openpyxl
import os

def generateDataBasicAuth():
    print(os.getcwd())
    wk = openpyxl.load_workbook(r'C:\Users\Patryk\PycharmProjects\The-Internet-Tests\Library\TestData.xlsx')
    sheet = wk['BasicAuthInvalidData']
    row = sheet.max_row
    list_data = []

    for i in range(2, row + 1):
        temporary_list = []
        username = sheet.cell(i, 1)
        password = sheet.cell(i, 2)
        temporary_list.insert(0, username.value)
        temporary_list.insert(1, password.value)
        list_data.insert(i - 2, temporary_list)

    return list_data
